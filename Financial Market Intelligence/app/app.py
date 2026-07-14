import os
import sys

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from datetime import datetime, timedelta

import joblib # For loading ML models
import numpy as np # For numerical operations
import pandas as pd # For data manipulation
import plotly.graph_objects as go # For creating interactive plots
import time # For adding delays
import streamlit as st # For creating the web application
import yfinance as yf # For fetching financial data
from src.data.fetch_data import fetch_all # For fetching data from the database
from src.processing.preprocess import preprocess # For preprocessing the data
from io import BytesIO # For handling binary data
from openpyxl.drawing.image import Image # For adding images to Excel files
from openpyxl.chart import BarChart3D, Reference # For creating 3D bar charts
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side # For styling Excel files
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table # For creating PDF reports
from reportlab.lib import colors # For defining colors
from reportlab.lib.styles import getSampleStyleSheet # For getting sample styles
import socket # For checking internet connection

def check_internet():
    try:
        socket.create_connection(("8.8.8.8", 53), timeout=3)
        return True
    except:
        return False

#  CONFIG
st.set_page_config(page_title="Financial Market Intelligence", page_icon="💰", layout="wide")

# GLOBAL RESPONSIVE CSS INJECTION
st.markdown("""
<style>
/* Streamlit standard block layout fixes for mobile */
[data-testid="stMetric"] {
    background-color: #1a1a1a;
    padding: 15px;
    border-radius: 10px;
}
/* Responsive Tables */
.element-container iframe {
    width: 100% !important;
}
div[data-testid="stHorizontalBlock"] {
    flex-wrap: wrap !important;
}
</style>
""", unsafe_allow_html=True)

# INTERNET CHECK
if st.query_params.get("restart"):
    st.query_params.clear()
    st.cache_data.clear()
    st.cache_resource.clear()
    st.rerun()

if st.query_params.get("offline"):
    st.query_params.clear()
    st.session_state.offline_mode = True
    st.rerun()

# Internet check (Responsive Overlay Added)
if not check_internet() and not st.session_state.get("offline_mode"):
    st.markdown("""
    <style>
    .offline-overlay {
        position: fixed;
        top: 0; left: 0;
        width: 100vw; height: 100vh;
        background: rgba(0,0,0,0.92);
        backdrop-filter: blur(6px);
        z-index: 9998;
    }

    .offline-card {
        position: fixed;
        top: 50%; left: 50%;
        transform: translate(-50%, -50%);
        background: #1a1a1a;
        padding: 30px 20px;
        border-radius: 20px;
        text-align: center;
        color: white;
        width: 90%;
        max-width: 420px;
        box-shadow: 0 0 60px rgba(0,0,0,0.8);
        z-index: 9999;
    }

    .offline-title {
        font-size: 22px;
        color: #ff4d4d;
        font-weight: bold;
        margin-bottom: 12px;
    }

    .offline-text {
        color: #aaa;
        font-size: 14px;
        margin-bottom: 25px;
    }

    .btn-row {
        display: flex;
        flex-direction: row;
        gap: 12px;
        justify-content: center;
        flex-wrap: wrap;
    }

    .btn {
        padding: 11px 22px;
        border-radius: 10px;
        font-weight: 600;
        font-size: 14px;
        cursor: pointer;
        text-decoration: none !important;
        transition: background 0.2s, transform 0.2s;
        flex: 1;
        min-width: 140px;
        text-align: center;
    }

    .btn-restart { background: #f2003c; color: white !important; }
    .btn-restart:hover { background: #ff1a1a; transform: scale(1.02); }
    .btn-offline { background: #444; color: white !important; }
    .btn-offline:hover { background: #555; transform: scale(1.02); }
    </style>

    <div class="offline-overlay"></div>
    <div class="offline-card">
        <div class="offline-title">⚠ Oops! Internet Not Connected</div>
        <div class="offline-text">
            Please check your internet connection or continue in offline mode.
        </div>
        <div class="btn-row">
            <a href="?restart=1" class="btn btn-restart">🗘 Restart App</a>
            <a href="?offline=1" class="btn btn-offline">⊘ Offline Mode</a>
        </div>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

#  STYLES Title (Responsive Padding)
st.markdown("""
<div style='padding: 5px; margin-bottom: 10px;'>
    <h1 style='color:white; border-left:6px solid #dd0000; padding-left:12px; font-weight:bold; font-size: calc(1.8rem + 1.2vw); margin-bottom: 5px;'>
    Financial Market Intelligence
    </h1>
    <p style='color:#aaa; margin-left:12px; font-size: calc(0.9rem + 0.2vw);'>
    Advanced Analytics for Precious Metals, Energy & Currency Markets
    </p>
</div>
""", unsafe_allow_html=True)

if st.session_state.get("offline_mode"):
    st.warning("⚠ You are in Offline Mode. Data may be outdated. Please turn on internet and restart app.")

# ADVANCED LOADING SCREEN
loading_placeholder = st.empty()
loading_placeholder.markdown("""
<style>
.loader-container {
    display:flex;
    flex-direction:column;
    justify-content:center;
    align-items:center;
    height:50vh;
    text-align:center;
    padding: 20px;
}
.loader {
    border: 6px solid #1a1a1a;
    border-top: 6px solid #4FC3F7;
    border-radius: 50%;
    width: 60px;
    height: 60px;
    animation: spin 1s linear infinite;
    margin-bottom:20px;
}
@keyframes spin {
    0% { transform: rotate(0deg); }
    100% { transform: rotate(360deg); }
}
.loading-text { font-size:18px; color:white; font-weight:500; text-shadow: 0 0 10px #4FC3F7; }
.loading-sub { font-size:13px; color:#888; margin-top:5px; }
</style>
<div class="loader-container">
    <div class="loader"></div>
    <div class="loading-text">Loading Market Intelligence...</div>
    <div class="loading-sub">Fetching Gold, Silver & Currency Data</div>
</div>
""", unsafe_allow_html=True)

#  STYLES TABLE (Responsive Overflow Wrapper)
st.markdown("""
<style>
.table-wrapper {
    overflow-x: auto;
    width: 100%;
    margin-bottom: 20px;
}
table {
    width: 100% !important;
    border-collapse: collapse;
    text-align: center;
    font-size: 14px;
}
th { background-color: #111; color: white; padding: 10px; text-align: center !important; font-size: 14px; }
td { padding: 10px; border-bottom: 1px solid #333; text-align: center !important; font-size: 13px; }
tr:hover { background-color: #1a1a1a; }
</style>
""", unsafe_allow_html=True)

def styled_subheader(text):
    st.markdown(f"""
    <h3 style='border-left: 5px solid #4FC3F7; padding-left: 10px; font-weight: 400; margin-top: 20px; font-size: calc(1.2rem + 0.4vw);'>
    {text}
    </h3><br>
    """, unsafe_allow_html=True)

@st.cache_data(ttl=3600)
def load_data():
    fetch_all()
    preprocess()
    try:
        df = pd.read_csv("data/processed/final_data.csv")
        if df.empty: raise ValueError("Empty dataset")
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        df = df.dropna(subset=['Date'])
        return df
    except Exception as e:
        st.error("Failed to load data. Please update data.")
        return pd.DataFrame()

with st.spinner(" Fetching market data..."):
    try:
        load_data()
    except Exception as e:
        st.warning(f"Data update skipped: {e}")

# LOAD MODELS
try:
    gold_model = joblib.load("Financial Market Intelligence/models/gold_model.pkl")
    silver_model = joblib.load("Financial Market Intelligence/models/silver_model.pkl")
    usd_model = joblib.load("Financial Market Intelligence/models/usd_model.pkl")
    eur_model = joblib.load("Financial Market Intelligence/models/eur_model.pkl")
    gbp_model = joblib.load("Financial Market Intelligence/models/gbp_model.pkl")
    platinum_model = joblib.load("Financial Market Intelligence/models/platinum_model.pkl")
    palladium_model = joblib.load("Financial Market Intelligence/models/palladium_model.pkl")
    copper_model = joblib.load("Financial Market Intelligence/models/copper_model.pkl")
    crude_model = joblib.load("Financial Market Intelligence/models/crude_oil_model.pkl")
    brent_model = joblib.load("Financial Market Intelligence/models/brent_oil_model.pkl")
    gas_model = joblib.load("Financial Market Intelligence/models/natural_gas_model.pkl")

    gold_metrics = joblib.load("Financial Market Intelligence/models/gold_metrics.pkl")
    silver_metrics = joblib.load("Financial Market Intelligence/models/silver_metrics.pkl")
    usd_metrics = joblib.load("Financial Market Intelligence/models/usd_metrics.pkl")
    eur_metrics = joblib.load("Financial Market Intelligence/models/eur_metrics.pkl")
    gbp_metrics = joblib.load("Financial Market Intelligence/models/gbp_metrics.pkl")
    platinum_metrics = joblib.load("Financial Market Intelligence/models/platinum_metrics.pkl")
    palladium_metrics = joblib.load("Financial Market Intelligence/models/palladium_metrics.pkl")
    copper_metrics = joblib.load("Financial Market Intelligence/models/copper_metrics.pkl")
    crude_metrics = joblib.load("Financial Market Intelligence/models/crude_oil_metrics.pkl")
    brent_metrics = joblib.load("Financial Market Intelligence/models/brent_oil_metrics.pkl")
    gas_metrics = joblib.load("Financial Market Intelligence/models/natural_gas_metrics.pkl")
except Exception as e:
    st.error(f"Model loading failed: {e}")
    st.stop()

df = load_data()
loading_placeholder.empty()

if not df.empty:
    st.caption(f"Latest available data: {df['Date'].max().date()}")
if df.empty or len(df) < 2:
    st.error("Not enough data available")
    st.stop()

latest = df.iloc[-1]
previous = df.iloc[-2]

@st.cache_data(ttl=300)
def load_usd_full():
    try:
        usd = yf.download("USDINR=X", period="1y", progress=False)
        if usd.empty: return pd.DataFrame()
        if isinstance(usd.columns, pd.MultiIndex): usd.columns = usd.columns.get_level_values(0)
        usd.reset_index(inplace=True)
        return usd
    except Exception: return pd.DataFrame()

usd = load_usd_full()

g24_change = latest['Gold_24K_1g'] - previous['Gold_24K_1g']
g22_change = latest['Gold_22K_1g'] - previous['Gold_22K_1g']
silver_change = latest['Silver_1g'] - previous['Silver_1g']
g18_change = latest['Gold_18K_1g'] - previous['Gold_18K_1g']
eur_change = latest['EUR_INR'] - previous['EUR_INR']
gbp_change = latest['GBP_INR'] - previous['GBP_INR']
platinum_change = latest['Platinum_1g'] - previous['Platinum_1g']
palladium_change = latest['Palladium_1g'] - previous['Palladium_1g']
copper_change = latest['Copper_1g'] - previous['Copper_1g']
crude_change = latest['Crude_Oil_INR_per_barrel'] - previous['Crude_Oil_INR_per_barrel']
brent_change = latest['Brent_Oil_INR_per_barrel'] - previous['Brent_Oil_INR_per_barrel']
gas_change = latest['Natural_Gas_INR'] - previous['Natural_Gas_INR']

usd_price, usd_prev = 0, 0
try:
    if usd is None or usd.empty: raise ValueError("USD data empty")
    usd = usd.dropna()
    if len(usd) >= 2:
        usd_price = float(usd['Close'].iloc[-1])
        usd_prev = float(usd['Close'].iloc[-2])
    elif len(usd) == 1:
        usd_price = float(usd['Close'].iloc[-1])
        usd_prev = usd_price
except Exception:
    if 'USD_INR' in df.columns and len(df) >= 2:
        usd_price = float(df['USD_INR'].iloc[-1])
        usd_prev = float(df['USD_INR'].iloc[-2])

usd_change = usd_price - usd_prev

def format_change(val):
    if val > 0: return f"<span style='color:#02ff99; font-weight:bold;'>▲ {abs(val):.2f}</span>"
    elif val < 0: return f"<span style='color:#ff4d4d; font-weight:bold;'>▼ {abs(val):.2f}</span>"
    return "<span style='color:gray;'>0</span>"

g24_html = format_change(g24_change)
g22_html = format_change(g22_change)
silver_html = format_change(silver_change)
usd_html = format_change(usd_change)
g18_html = format_change(g18_change)
eur_html = format_change(eur_change)
gbp_html = format_change(gbp_change)
platinum_html = format_change(platinum_change)
palladium_html = format_change(palladium_change)
copper_html = format_change(copper_change)
crude_html = format_change(crude_change)
brent_html = format_change(brent_change)
gas_html = format_change(gas_change)

ticker_text = f"Gold 24K: ₹ {latest['Gold_24K_1g']:.2f} ({g24_html}) | Gold 22K: ₹ {latest['Gold_22K_1g']:.2f} ({g22_html}) | Gold 18K: ₹ {latest['Gold_18K_1g']:.2f} ({g18_html}) | Silver: ₹ {latest['Silver_1g']:.2f} ({silver_html}) | Platinum: ₹ {latest['Platinum_1g']:.2f} ({platinum_html}) | Palladium: ₹ {latest['Palladium_1g']:.2f} ({palladium_html}) | Copper: ₹ {latest['Copper_1g']:.2f} ({copper_html}) | USD: ₹ {usd_price:.2f} ({usd_html}) | EUR: ₹ {latest['EUR_INR']:.2f} ({eur_html}) | GBP: ₹ {latest['GBP_INR']:.2f} ({gbp_html}) | Crude Oil: ₹ {latest['Crude_Oil_INR_per_barrel']:.2f} ({crude_html}) | Brent Oil: ₹ {latest['Brent_Oil_INR_per_barrel']:.2f} ({brent_change}) | Natural Gas: ₹ {latest['Natural_Gas_INR']:.2f} ({gas_html}) | "

st.markdown(f"""
<style>
.ticker-container {{ width: 100%; overflow: hidden; background: #0e1117; padding: 10px 0; }}
.ticker-track {{ display: flex; width: max-content; animation: scroll 30s linear infinite; }}
.ticker-item {{ white-space: nowrap; color: white; font-size: 15px; padding-right: 50px; }}
@keyframes scroll {{ 0% {{ transform: translateX(0); }} 100% {{ transform: translateX(-50%); }} }}
</style>
<div class="ticker-container"><div class="ticker-track"><div class="ticker-item">{ticker_text}</div><div class="ticker-item">{ticker_text}</div></div></div>
""", unsafe_allow_html=True)

# Helper Input Data Creators
def create_gold_input(df):
    lag1 = df['Gold_24K_1g'].iloc[-1]
    lag2 = df['Gold_24K_1g'].iloc[-2] if len(df) > 1 else lag1
    lag3 = df['Gold_24K_1g'].iloc[-3] if len(df) > 2 else lag2
    ma7, ma30 = df['Gold_24K_1g'].tail(7).mean(), df['Gold_24K_1g'].tail(30).mean()
    usd = df['USD_INR'].iloc[-1]
    usd_change = df['USD_INR'].pct_change().iloc[-1] if len(df) > 1 else 0
    g22, g18 = df['Gold_22K_1g'].iloc[-1], df['Gold_18K_1g'].iloc[-1]
    return pd.DataFrame([[lag1, lag2, lag3, ma7, ma30, usd, usd_change, g22, g18, g18/lag1, df['Date'].iloc[-1].dayofweek]], columns=['Lag_1','Lag_2','Lag_3','MA_7','MA_30','USD_INR','USD_Change','Gold_22K_1g','Gold_18K_1g','Gold_18K_Ratio','DayOfWeek'])

def create_silver_input(df):
    lag1 = df['Silver_1g'].iloc[-1]
    lag2 = df['Silver_1g'].iloc[-2] if len(df) > 1 else lag1
    lag3 = df['Silver_1g'].iloc[-3] if len(df) > 2 else lag2
    return pd.DataFrame([[lag1, lag2, lag3, df['Silver_1g'].tail(3).mean(), df['Silver_1g'].tail(7).mean(), df['USD_INR'].iloc[-1], df['USD_INR'].pct_change().iloc[-1] if len(df) > 1 else 0]], columns=['Lag_1','Lag_2','Lag_3','MA_3','MA_7','USD_INR','USD_Change'])

def create_usd_input(df):
    return pd.DataFrame([[df['USD_INR'].iloc[-1], df['USD_INR'].iloc[-2], df['USD_INR'].iloc[-3], df['USD_INR'].tail(3).mean(), df['USD_INR'].tail(7).mean()]], columns=['Lag_1','Lag_2','Lag_3','MA_3','MA_7'])

def create_eur_input(df):
    return pd.DataFrame([[df['EUR_INR'].iloc[-1], df['EUR_INR'].iloc[-2], df['EUR_INR'].iloc[-3], df['EUR_INR'].tail(3).mean(), df['EUR_INR'].tail(7).mean()]], columns=['Lag_1','Lag_2','Lag_3','MA_3','MA_7'])

def create_gbp_input(df):
    return pd.DataFrame([[df['GBP_INR'].iloc[-1], df['GBP_INR'].iloc[-2], df['GBP_INR'].iloc[-3], df['GBP_INR'].tail(3).mean(), df['GBP_INR'].tail(7).mean()]], columns=['Lag_1','Lag_2','Lag_3','MA_3','MA_7'])

def create_metal_input(df, col):
    return pd.DataFrame([[df[col].iloc[-1], df[col].iloc[-2] if len(df) > 1 else df[col].iloc[-1], df[col].iloc[-3] if len(df) > 2 else df[col].iloc[-1], df[col].tail(3).mean(), df[col].tail(7).mean(), df['USD_INR'].iloc[-1], df['USD_INR'].pct_change().iloc[-1] if len(df) > 1 else 0]], columns=['Lag_1','Lag_2','Lag_3','MA_3','MA_7','USD_INR','USD_Change'])

def get_prediction(df, metal):
    if metal == "Gold_24K": return gold_model.predict(create_gold_input(df))[0]
    elif metal == "Gold_22K": return gold_model.predict(create_gold_input(df))[0] * (22/24)
    elif metal == "Gold_18K": return gold_model.predict(create_gold_input(df))[0] * (18/24)
    elif metal == "Silver": return silver_model.predict(create_silver_input(df))[0]
    elif metal == "USD": return usd_model.predict(create_usd_input(df))[0]
    elif metal == "EUR": return eur_model.predict(create_eur_input(df))[0]
    elif metal == "GBP": return gbp_model.predict(create_gbp_input(df))[0]
    elif metal == "Platinum": return platinum_model.predict(create_metal_input(df, 'Platinum_1g'))[0]
    elif metal == "Palladium": return palladium_model.predict(create_metal_input(df, 'Palladium_1g'))[0]
    elif metal == "Copper": return copper_model.predict(create_metal_input(df, 'Copper_1g'))[0]
    elif metal == "Crude_Oil": return crude_model.predict(create_metal_input(df, 'Crude_Oil_INR_per_barrel'))[0]
    elif metal == "Brent_Oil": return brent_model.predict(create_metal_input(df, 'Brent_Oil_INR_per_barrel'))[0]
    elif metal == "Natural_Gas": return gas_model.predict(create_metal_input(df, 'Natural_Gas_INR'))[0]
    return None

# Responsive Card Template (Font scales dynamically)
def render_card(title, value, change=None):
    arrow, color, change_html = "", "#aaa", ""
    if change is not None:
        arrow = "▲" if change > 0 else "▼"
        color = "#00ff99" if change > 0 else "#ff4d4d"
        change_html = f"<span style='color:{color}; font-size:calc(12px + 0.3vw); margin-left:6px;'>({arrow} {abs(change):.2f})</span>"
    
    st.markdown(f"""
        <div style="display:flex; flex-direction:column; gap:4px; background:#1e1e1e; padding:12px; border-radius:10px; border: 1px solid #333; margin-bottom:10px;">
            <span style="color:#aaa; font-size:calc(11px + 0.2vw);">{title}</span>
            <span style="font-size:calc(18px + 0.6vw); font-weight:600; color:white; white-space:nowrap;">₹ {value:.2f} {change_html}</span>
        </div>
        """, unsafe_allow_html=True)

def show_section(asset):
    styled_subheader(f"{asset} Overview")
    max_date = df['Date'].max().date()
    
    asset_type = "metal" if asset in ["Gold_24K", "Gold_22K", "Gold_18K", "Silver", "Platinum", "Palladium", "Copper"] else ("currency" if asset in ["USD", "EUR", "GBP"] else "energy")

    # Responsive Input Block
    if asset_type == "currency":
        selected_date = st.date_input("Choose Date", value=max_date, min_value=df['Date'].min().date(), max_value=datetime.now().date() + timedelta(days=7), key=f"{asset}_date")
        multiplier, base_col = 1, f"{asset}_INR"
        selected_weight = "unit"
    else:
        colA, colB = st.columns([1, 1])
        with colA:
            selected_date = st.date_input("Choose Date", value=max_date, min_value=df['Date'].min().date(), max_value=datetime.now().date() + timedelta(days=7), key=f"{asset}_date")
        with colB:
            if asset_type == "metal":
                selected_weight = st.selectbox("Select Weight", ["1g", "10g", "100g", "1kg"], index=0, key=f"{asset}_weight")
                multiplier = {"1g":1, "10g":10, "100g":100, "1kg":1000}[selected_weight]
                base_col = f"{asset}_1g"
            else:
                lbl = "Select Volume"
                opts = ["1 Unit", "10 Unit", "100 Unit", "1000 Unit"] if asset == "Natural_Gas" else ["1 Barrel", "10 Barrel", "100 Barrel", "1000 Barrel"]
                selected_weight = st.selectbox(lbl, opts, index=0, key=f"{asset}_barrel")
                multiplier = {opts[0]:1, opts[1]:10, opts[2]:100, opts[3]:1000}[selected_weight]
                base_col = "Natural_Gas_INR" if asset == "Natural_Gas" else f"{asset}_INR_per_barrel"

    if base_col not in df.columns:
        st.error(f"{asset} data not available")
        return

    filtered_df = df[df['Date'].dt.date == selected_date]
    if not filtered_df.empty:
        selected_row = filtered_df.iloc[0]
        today_price = float(selected_row[base_col]) * multiplier
        prev_df = df[df['Date'] < pd.to_datetime(selected_date)]
        yesterday_price = (float(prev_df.iloc[-1][base_col]) if not prev_df.empty else float(selected_row[base_col])) * multiplier
    else:
        st.warning("Future date selected — showing prediction")
        temp_df = df.copy()
        for _ in range((selected_date - max_date).days):
            next_pred = get_prediction(temp_df, asset)
            new_row = temp_df.iloc[-1].copy()
            new_row['Date'] += timedelta(days=1)
            new_row[base_col] = next_pred
            temp_df = pd.concat([temp_df, pd.DataFrame([new_row])], ignore_index=True)
        selected_row = temp_df.iloc[-1]
        today_price = float(selected_row[base_col]) * multiplier
        yesterday_price = float(temp_df.iloc[-2][base_col]) * multiplier

    change = today_price - yesterday_price

    # 4 metrics wrap grids neatly on mobile using automatic flex containers or sequential render
    m_cols = st.columns([1, 1, 1, 1])
    with m_cols[0]: render_card("📅 Selected Day", today_price, change)
    with m_cols[1]: render_card("🗓️ Previous Day", yesterday_price)
    with m_cols[2]: render_card("📈 Highest", df[base_col].max() * multiplier)
    with m_cols[3]: render_card("📉 Lowest", df[base_col].min() * multiplier)

    try:
        prediction = get_prediction(df, asset) * multiplier
        st.markdown("<br>", unsafe_allow_html=True)
        render_card("🎯 Predicted Next Day", prediction, prediction - today_price)
    except Exception as e:
        st.error(f"Prediction error: {e}")

    # Responsive Tables Rendering
    if asset_type in ["metal", "energy"]:
        styled_subheader("Price Table")
        weights = ["1g", "10g", "100g", "1kg"] if asset_type == "metal" else (["1 Unit", "10 Unit", "100 Unit", "1000 Unit"] if asset == "Natural_Gas" else ["1 Barrel", "10 Barrel", "100 Barrel", "1000 Barrel"])
        rows = []
        for w in weights:
            mul = {"1g":1, "10g":10, "100g":100, "1kg":1000, "1 Unit":1, "10 Unit":10, "100 Unit":100, "1000 Unit":1000, "1 Barrel":1, "10 Barrel":10, "100 Barrel":100, "1000 Barrel":1000}[w]
            t = float(selected_row[base_col]) * mul
            prev_df = df[df['Date'] < pd.to_datetime(selected_date)]
            y = (float(prev_df.iloc[-1][base_col]) if not prev_df.empty else float(selected_row[base_col])) * mul
            c_html = f"<span style='color:#02ff99'>▲ ₹{abs(t-y):,.2f}</span>" if (t-y) > 0 else f"<span style='color:#ff4d4d'>▼ ₹{abs(t-y):,.2f}</span>"
            rows.append({"Unit": w, "Today": f"₹{t:,.2f}", "Yesterday": f"₹{y:,.2f}", "Change": c_html})
        
        st.markdown(f'<div class="table-wrapper">{pd.DataFrame(rows).to_html(escape=False, index=False)}</div>', unsafe_allow_html=True)

    # 7-day prediction table
    styled_subheader("7 Day Prediction")
    future_preds, future_dates, temp_df = [], [], df.copy()
    start_date = pd.to_datetime(selected_date)
    for i in range(7):
        next_pred = get_prediction(temp_df, asset)
        future_preds.append(next_pred * multiplier)
        future_dates.append(start_date + timedelta(days=i+1))
        new_row = temp_df.iloc[-1].copy()
        new_row['Date'] += timedelta(days=1)
        new_row[base_col] = next_pred
        temp_df = pd.concat([temp_df, pd.DataFrame([new_row])], ignore_index=True)

    pred_rows = []
    for i in range(7):
        prev_val = today_price if i == 0 else future_preds[i-1]
        diff = future_preds[i] - prev_val
        c_html = f"<span style='color:#02ff99'>▲ ₹{abs(diff):,.2f}</span>" if diff > 0 else f"<span style='color:#ff4d4d'>▼ ₹{abs(diff):,.2f}</span>"
        pred_rows.append({"Date": future_dates[i].date().strftime('%d-%m-%Y'), "Predicted Price": f"₹{future_preds[i]:,.2f}", "Change": c_html})
    
    st.markdown(f'<div class="table-wrapper">{pd.DataFrame(pred_rows).to_html(escape=False, index=False)}</div>', unsafe_allow_html=True)

    # Chart Integration (Responsive Plotly)
    styled_subheader("Price Trend")
    btn_cols = st.columns(6)
    ranges = ["1W", "1M", "3M", "6M", "1Y", "ALL"]
    selected_range = "ALL"
    for i, r in enumerate(ranges):
        if btn_cols[i].button(r, key=f"{asset}_btn_{r}"): selected_range = r

    dff = df.tail(7) if selected_range == "1W" else (df.tail(30) if selected_range == "1M" else (df.tail(90) if selected_range == "3M" else (df.tail(180) if selected_range == "6M" else (df.tail(365) if selected_range == "1Y" else df.copy()))))
    
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=dff['Date'], y=dff[base_col] * multiplier, mode='lines', line=dict(color='#00ff99', width=3), name='Actual'))
    fig.add_trace(go.Scatter(x=future_dates, y=future_preds, mode='lines+markers', line=dict(color='#ffaa00', width=3, dash='dot'), name='Prediction'))
    fig.update_layout(template="plotly_dark", hovermode="x unified", yaxis_title=f"Price ({selected_weight})", margin=dict(l=10, r=10, t=30, b=10))
    st.plotly_chart(fig, use_container_width=True)

# Tabs
tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8, tab9, tab10, tab11, tab12, tab13, tab14 = st.tabs([
    "🪙 Gold 24K", "🧈 Gold 22K", "🟡 Gold 18K", "🔘 Silver", "🪨 Platinum", "☢️ Palladium", "🥮 Copper", "💱 USD", "💶 EUR", "💷 GBP", "⛽ Crude Oil", "🛢️ Brent Oil", "♨️ Natural Gas", "🌟 Model Performance"
])

with tab1: show_section("Gold_24K")
with tab2: show_section("Gold_22K")
with tab3: show_section("Gold_18K")
with tab4: show_section("Silver")
with tab5: show_section("Platinum")
with tab6: show_section("Palladium")
with tab7: show_section("Copper")
with tab8: show_section("USD")
with tab9: show_section("EUR")
with tab10: show_section("GBP")
with tab11: show_section("Crude_Oil")
with tab12: show_section("Brent_Oil")
with tab13: show_section("Natural_Gas")

with tab14:
    styled_subheader("🎯 Model Performance (All Assets)")
    rows = []
    def add_row(name, metrics):
        if metrics: rows.append({"Model": name, "MAE": f"{metrics.get('MAE', 0):.2f}", "RMSE": f"{metrics.get('RMSE', 0):.2f}", "R² Score": f"{metrics.get('R2', 0):.4f}"})
    
    try:
        add_row("Gold", gold_metrics)
        add_row("Silver", silver_metrics)
        add_row("Platinum", platinum_metrics)
        add_row("Palladium", palladium_metrics)
        add_row("Copper", copper_metrics)
        add_row("USD", usd_metrics)
        add_row("EUR", eur_metrics)
        add_row("GBP", gbp_metrics)
        add_row("Crude Oil", crude_metrics)
        add_row("Brent Oil", brent_metrics)
        add_row("Natural Gas", gas_metrics)
    except: pass

    table_df = pd.DataFrame(rows)
    r2_values = [float(r["R² Score"]) for r in rows if r["R² Score"] != "-"]
    if r2_values: table_df.loc[len(table_df)] = {"Model": "Model R² Accuracy (Average)", "MAE": "-", "RMSE": "-", "R² Score": f"{sum(r2_values)/len(r2_values):.4f}"}
    
    st.markdown(f'<div class="table-wrapper">{table_df.to_html(index=False)}</div>', unsafe_allow_html=True)
    
    st.markdown("---")
    styled_subheader("📘 Calculation Metrics Explained")
    
    # 3-column wrap setup for descriptions
    c_m1, c_m2, c_m3 = st.columns([1, 1, 1])
    with c_m1: st.markdown("### 📏 MAE\n- Average error size.\n- Lower = Better.")
    with c_m2: st.markdown("### 📐 RMSE\n- Penalizes large errors.\n- Sensitive to outliers.")
    with c_m3: st.markdown("### 🎯 R² Score\n- Model accuracy fit.\n- 1 = Perfect.")

    st.markdown("---")
    styled_subheader("🎯 Model Performance Flow")
    plot_df = table_df[table_df["Model"] != "Model R² Accuracy (Average)"].copy()
    plot_df["R2_float"] = plot_df["R² Score"].astype(float)
    
    fig_perf = go.Figure()
    fig_perf.add_trace(go.Scatter(x=plot_df["Model"], y=plot_df["R2_float"], mode='lines+markers', marker=dict(symbol='circle', size=10, color='#00a600'), line=dict(color='#00a600', width=3)))
    fig_perf.update_layout(xaxis=dict(title="Models"), yaxis=dict(title="R² Score"), paper_bgcolor="#0E1117", plot_bgcolor="#0E1117", font=dict(color="white"), margin=dict(l=10, r=10, t=30, b=10))
    st.plotly_chart(fig_perf, use_container_width=True)

# Dialog Popup Setup
@st.dialog(" ", width="large")
def show_popup(asset, latest):
    icon = {"Gold": "🧈", "Silver": "🔘", "Platinum": "🪨", "Palladium": "☢️", "Copper": "🥮"}.get(asset, "⚖️")
    st.markdown(f"## {icon} {asset} Calculator")
    premium_calculator(asset, latest)

def generate_pdf_report(asset, purity, price, qty_grams, base, making_amt, gst_amt, total):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer)
    styles = getSampleStyleSheet()
    elements = [Paragraph(f"<b>{asset} Investment Report</b>", styles['Title']), Spacer(1, 10), Paragraph("Financial Market Intelligence", styles['Normal']), Spacer(1, 20)]
    table = Table([["Field", "Value"], ["Asset", asset], ["Purity", purity], ["Rate", f"₹{price}"], ["Weight", f"{qty_grams} g"], ["Total", f"₹{total:,}"]])
    table.setStyle([("BACKGROUND", (0,0), (-1,0), colors.green), ("TEXTCOLOR", (0,0), (-1,0), colors.white), ("GRID", (0,0), (-1,-1), 1, colors.black)])
    elements.extend([table, Spacer(1, 20), Paragraph("<b>⚠ This is an estimated ML-based report. Market risk involved.</b>", styles['Normal'])])
    doc.build(elements)
    return buffer.getvalue()

def generate_excel_report(asset, purity, price, qty_grams, base, making_amt, gst_amt, total, df):
    future_dates, future_prices, temp_df = [], [], df.copy()
    metal_key = asset if asset != "Gold" else f"Gold_{purity}"
    for i in range(7):
        pred = get_prediction(temp_df, metal_key)
        future_prices.append(int(pred))
        next_date = temp_df['Date'].iloc[-1] + timedelta(days=1)
        future_dates.append(next_date.strftime("%d-%m-%Y"))
        new_row = temp_df.iloc[-1].copy()
        new_row['Date'] = next_date
        if asset == "Gold": new_row[f"{metal_key}_1g"] = pred
        else: new_row[{"Silver": "Silver_1g", "Platinum": "Platinum_1g", "Palladium": "Palladium_1g", "Copper": "Copper_1g"}[asset]] = pred
        temp_df = pd.concat([temp_df, pd.DataFrame([new_row])], ignore_index=True)

    calc_data = [["Asset", asset], ["Purity", purity if asset == "Gold" else "-"], ["Rate (₹/g)", int(price)], ["Weight (g)", round(qty_grams, 3)], ["Base Value", int(base)], ["Making Charges", int(making_amt)], ["GST (+3%)", int(gst_amt)], ["Total Amount", int(total)]]
    pred_df = pd.DataFrame({"Date": future_dates, "Predicted Price (₹/g)": future_prices})
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pd.DataFrame(calc_data, columns=["Field", "Value"]).to_excel(writer, index=False, startrow=5, startcol=0)
        pred_df.to_excel(writer, index=False, startrow=5, startcol=3)
        ws = writer.sheets['Sheet1']
        for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=13):
            for cell in row: cell.fill = PatternFill("solid", fgColor="FFFFFF")
        for r in range(1, 5):
            for c in range(1, 14): ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="1E3A5F" if r < 4 else "C9A84C")
        ws.merge_cells("A2:M2")
        ws["A2"] = f"{asset} Investment Report"
        ws["A2"].font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.sheet_view.showGridLines = False
    return output.getvalue()

def premium_calculator(asset, latest):
    price_map = {"Gold": latest["Gold_24K_1g"], "Silver": latest["Silver_1g"], "Platinum": latest["Platinum_1g"], "Palladium": latest["Palladium_1g"], "Copper": latest["Copper_1g"]}
    prefix = f"popup_{asset}"

    st.session_state.setdefault(f"{prefix}_mode", "weight")
    st.session_state.setdefault(f"{prefix}_purity", "24K")
    st.session_state.setdefault(f"{prefix}_val", 10)
    st.session_state.setdefault(f"{prefix}_val_amt", 10000)
    st.session_state.setdefault(f"{prefix}_unit", "Gram")

    # Flex block: Converts columns to stacked layout smoothly on mobile screens
    left, right = st.columns([1, 1])

    with left:
        c1, c2 = st.columns(2)
        if c1.button("Calculate By Weight", key=f"{prefix}_w"): st.session_state[f"{prefix}_mode"] = "weight"
        if c2.button("Calculate By Amount", key=f"{prefix}_a"): st.session_state[f"{prefix}_mode"] = "amount"

        mode, purity = st.session_state[f"{prefix}_mode"], st.session_state[f"{prefix}_purity"]
        if asset == "Gold":
            st.markdown("### Gold Purity")
            p1, p2, p3 = st.columns(3)
            if p1.button("24K"): st.session_state[f"{prefix}_purity"] = "24K"
            if p2.button("22K"): st.session_state[f"{prefix}_purity"] = "22K"
            if p3.button("18K"): st.session_state[f"{prefix}_purity"] = "18K"
            purity = st.session_state[f"{prefix}_purity"]

        price = price_map[asset] * {"24K":1, "22K":22/24, "18K":18/24}.get(purity, 1) if asset == "Gold" else price_map[asset]
        unit_map = {"Gram": 1, "Sovereign / Pavan": 8, "Tola": 11.664, "KG": 1000}

        if mode == "weight":
            q1, q2 = st.columns([2, 1])
            with q2: unit = st.selectbox("Unit", ["Gram", "Sovereign / Pavan", "Tola", "KG"], key=f"{prefix}_unit")
            max_val = 1000 if unit == "Gram" else 100
            
            st.session_state[f"{prefix}_val"] = min(st.session_state[f"{prefix}_val"], max_val)
            with q1: qty = st.number_input("Quantity", 1, max_val, value=int(st.session_state[f"{prefix}_val"]), key=f"{prefix}_input")
            
            st.session_state[f"{prefix}_val"] = qty
            final_weight = float(qty)
            qty_grams = qty * unit_map[unit]
            making = st.slider("Making Charge (%)", 0, 50, 10, key=f"{prefix}_mkm")
            gst = st.checkbox("Include GST (3%)", True, key=f"{prefix}_gstk")
            base = qty_grams * price
        else:
            amt = st.number_input("Amount (₹)", 0, 1000000, value=int(st.session_state[f"{prefix}_val_amt"]), key=f"{prefix}_amt_input")
            st.session_state[f"{prefix}_val_amt"] = amt
            making = st.slider("Making Charge (%)", 0, 50, 10, key=f"{prefix}_mka")
            gst = st.checkbox("Include GST (3%)", True, key=f"{prefix}_gsta")
            base = float(amt)
            qty_grams = base / price if price > 0 else 0
            final_weight = float(f"{qty_grams:.3f}")

        making_amt = base * (making/100)
        subtotal = base + making_amt
        gst_amt = subtotal * 0.03 if gst else 0
        total = subtotal + gst_amt

        st.markdown("<br>", unsafe_allow_html=True)
        st.download_button(label="📥 Download Full Report", data=generate_excel_report(asset, purity, price, qty_grams, base, making_amt, gst_amt, total, df), file_name=f"{asset}_full_report.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    with right:
        rate_label = f"{purity} Current Rate" if asset == "Gold" else f"{asset} Current Rate"
        title = f"{asset} Weight" if mode == "amount" else "Total Amount"
        text = f"{final_weight:.3f} g" if mode == "amount" else f"₹ {int(total):,}"
        
        # Responsive Calculation Grid Cards using Flexbox-wrap
        st.markdown(f"""
        <style>
        .calc-wrapper {{ display: flex; flex-direction: column; gap: 10px; color: white; margin-top: 10px; }}
        .row-item {{ display: flex; justify-content: space-between; padding: 6px 0; border-bottom: 1px solid #333; font-size: 14px; }}
        </style>
        <div class="calc-wrapper">
            <div style="background:#15803d; padding:15px; border-radius:12px; color:white">
                <div style="font-size:16px; opacity:0.9;">{title}</div>
                <div style="font-size:calc(22px + 1vw); font-weight:800;">{text}</div>
            </div>
            <div style="background:rgba(255,255,255,0.06); padding:10px; border-radius:8px;">
                <div style="font-size:12px; color:#aaa;">{rate_label}</div>
                <div style="font-size:16px; font-weight:700;">₹{int(price):,}/gram</div>
            </div>
            <div style="margin-top:10px;">
                <h4 style="color:white; font-size: 15px; margin-bottom:8px;">Calculation Breakdown</h4>
                <div class="row-item"><span>Weight</span><b>{final_weight:.3f} g</b></div>
                <div class="row-item"><span>Base Value</span><b>₹ {int(base):,}</b></div>
                <div class="row-item"><span>Making Charges</span><b>₹ {int(making_amt):,}</b></div>
                <div class="row-item"><span>Subtotal</span><b>₹ {int(subtotal):,}</b></div>
                <div class="row-item"><span>GST (3%)</span><b>₹ {int(gst_amt):,}</b></div>
                <div class="row-item" style="background:#166534; padding:8px; border-radius:6px; margin-top:5px; border:none;">
                    <span>Total Amount</span><b>₹ {int(total):,}</b>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

# Sidebar (Dynamic full-width buttons for mobile)
with st.sidebar:
    st.markdown("""
    <style>
    [data-testid="stSidebar"] [data-te stid="stElementContainer"] { width: 100% !important; }
    [data-testid="stSidebar"] button {
        width: 100% !important; height: 50px !important; border-radius: 12px !important;
        font-size: 15px !important; font-weight: 600 !important; text-align: left !important;
        padding-left: 15px !important; display: flex !important; align-items: center !important;
    }
    </style>
    <h2 style='color:white; border-left:6px solid #00a600; padding-left:12px; font-weight:bold; font-size: 20px;'>⚖️ Metal Calculator</h2>
    <div style='height:10px;'></div>
    """, unsafe_allow_html=True)

    if st.button("Gold"): st.session_state.popup = "Gold"
    if st.button("Silver"): st.session_state.popup = "Silver"
    if st.button("Platinum"): st.session_state.popup = "Platinum"
    if st.button("Palladium"): st.session_state.popup = "Palladium"
    if st.button("Copper"): st.session_state.popup = "Copper"

    st.markdown("""
    <br><h2 style='color:white; border-left:6px solid #00a600; padding-left:12px; font-weight:bold; font-size: 20px;'>🛠️ Settings</h2>
    <div style='height:10px;'></div>
    <style>
    .btn-restart { display:block; text-align:center; padding:10px; background:#f2003c; color:white !important; border-radius:10px; text-decoration:none !important; font-weight:600; font-size:14px;}
    .btn-restart:hover { background: #ff1a1a; }
    </style>
    <a href="?restart=1" class="btn btn-restart">🗘 Restart App</a>
    """, unsafe_allow_html=True)

if "popup" not in st.session_state: st.session_state.popup = None
if st.session_state.popup:
    asset_to_show = st.session_state.popup
    st.session_state.popup = None
    show_popup(asset_to_show, latest)

st.markdown("---")
st.caption("© 2026 • Developed by Daksh Vasani | Advanced Analytics • Machine Learning • Financial Insights")
