import os
import sys

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from datetime import datetime, timedelta

import joblib # For loading ML models
import numpy as np # For numerical operations
import pandas as pd # For data manipulation
import plotly.graph_objects as go # For creating interactive plots
import time # For adding delays
import streamlit as st # For creating the web application
import yfinance as yf # For fetching financial data
import plotly.graph_objects as go # For creating interactive plots
from src.data.fetch_data import fetch_all # For fetching data from the database
from src.processing.preprocess import preprocess # For preprocessing the data
from io import BytesIO # For handling binary data
from openpyxl.drawing.image import Image # For adding images to Excel files
from openpyxl.chart import BarChart3D, Reference # For creating 3D bar charts
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side # For styling Excel files
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table # For creating PDF reports
from reportlab.lib import colors # For defining colors
from reportlab.lib.styles import getSampleStyleSheet # For getting sample styles

#  CONFIG
st.set_page_config(page_title="Financial Market Intelligence", page_icon="💰", layout="wide")
#  STYLES Title
st.markdown("""
<h1 style='
    color:white;
    border-left:6px solid #dd0000;
    padding-left:12px;
    font-weight:bold;
'>
Financial Market Intelligence
</h1>
<p style='color:#aaa; margin-left:12px;'>
Advanced Analytics for Precious Metals, Energy & Currency Markets
</p>
""", unsafe_allow_html=True)

# ADVANCED LOADING SCREEN
loading_placeholder = st.empty()

loading_placeholder.markdown("""
<style>
.loader-container {
    display:flex;
    flex-direction:column;
    justify-content:center;
    align-items:center;
    height:60vh;
    text-align:center;
}

.loader {
    border: 6px solid #1a1a1a;
    border-top: 6px solid #4FC3F7;
    border-radius: 50%;
    width: 70px;
    height: 70px;
    animation: spin 1s linear infinite;
    margin-bottom:20px;
}

@keyframes spin {
    0% { transform: rotate(0deg); }
    100% { transform: rotate(360deg); }
}

.loading-text {
    font-size:20px;
    color:white;
    font-weight:500;
    text-shadow: 0 0 10px #4FC3F7;
}

.loading-sub {
    font-size:14px;
    color:#888;
    margin-top:5px;
}
</style>

<div class="loader-container">
    <div class="loader"></div>
    <div class="loading-text">Loading Market Intelligence...</div>
    <div class="loading-sub">Fetching Gold, Silver & Currency Data</div>
</div>
""", unsafe_allow_html=True)

#  STYLES TABLE
st.markdown("""
<style>
table {
    width: 100% !important;
    border-collapse: collapse;
    text-align: center;
    font-size: 16px;
}
th {
    background-color: #111;
    color: white;
    padding: 14px;
    text-align: center !important;
}
td {
    padding: 12px;
    border-bottom: 1px solid #333;
    text-align: center !important;
}
tr:hover {
    background-color: #1a1a1a;
}
</style>
""", unsafe_allow_html=True)

#  STYLES SUBHEADER
def styled_subheader(text):
    st.markdown(f"""
    <h3 style='
        border-left: 5px solid #4FC3F7;
        padding-left: 10px;
        font-weight: 400;
        margin-top: 20px;
    '>
    {text}
    </h3><br>
    """, unsafe_allow_html=True)


@st.cache_data(ttl=3600)   # 1 hour cache
def load_data():
    fetch_all()
    preprocess()

    try:
        df = pd.read_csv("data/processed/final_data.csv")

        if df.empty:
            raise ValueError("Empty dataset")

        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        df = df.dropna(subset=['Date'])

        return df

    except Exception as e:
        st.error("Failed to load data. Please update data.")
        return pd.DataFrame()
#  AUTO REFRESH
with st.spinner(" Fetching market data..."):
    try:
        load_data()
    except Exception as e:
        st.warning(f"Data update skipped: {e}")

# LOAD MODEL
try:
    # CORE MODELS
    gold_model = joblib.load("models/gold_model.pkl")
    silver_model = joblib.load("models/silver_model.pkl")
    usd_model = joblib.load("models/usd_model.pkl")

    # CURRENCY MODELS
    eur_model = joblib.load("models/eur_model.pkl")
    gbp_model = joblib.load("models/gbp_model.pkl")

    # METALS
    platinum_model = joblib.load("models/platinum_model.pkl")
    palladium_model = joblib.load("models/palladium_model.pkl")
    copper_model = joblib.load("models/copper_model.pkl")

    # ENERGY
    crude_model = joblib.load("models/crude_oil_model.pkl")
    brent_model = joblib.load("models/brent_oil_model.pkl")
    gas_model = joblib.load("models/natural_gas_model.pkl")

    # METRICS
    gold_metrics = joblib.load("models/gold_metrics.pkl")
    silver_metrics = joblib.load("models/silver_metrics.pkl")
    usd_metrics = joblib.load("models/usd_metrics.pkl")
    eur_metrics = joblib.load("models/eur_metrics.pkl")
    gbp_metrics = joblib.load("models/gbp_metrics.pkl")

    platinum_metrics = joblib.load("models/platinum_metrics.pkl")
    palladium_metrics = joblib.load("models/palladium_metrics.pkl")
    copper_metrics = joblib.load("models/copper_metrics.pkl")

    crude_metrics = joblib.load("models/crude_oil_metrics.pkl")
    brent_metrics = joblib.load("models/brent_oil_metrics.pkl")
    gas_metrics = joblib.load("models/natural_gas_metrics.pkl")

except Exception as e:
    st.error(f"Model loading failed: {e}")
    st.stop()
df = load_data()
loading_placeholder.empty()
if not df.empty:
    st.caption(f"Latest available data: {df['Date'].max().date()}")
if df.empty or len(df) < 2:
    st.error("Not enough data available")
    st.stop()

latest = df.iloc[-1]
previous = df.iloc[-2]

today_date = latest['Date'].date()
yesterday_date = previous['Date'].date()

#  USD DATA 
@st.cache_data(ttl=300)
def load_usd_full():
    usd = yf.download("USDINR=X", period="1y")
    if isinstance(usd.columns, pd.MultiIndex):
        usd.columns = usd.columns.get_level_values(0)
    usd.reset_index(inplace=True)
    return usd


usd = load_usd_full()


#  CHANGE CALCULATION 
g24_change = latest['Gold_24K_1g'] - previous['Gold_24K_1g']
g22_change = latest['Gold_22K_1g'] - previous['Gold_22K_1g']
silver_change = latest['Silver_1g'] - previous['Silver_1g']
g18_change = latest['Gold_18K_1g'] - previous['Gold_18K_1g']
eur_change = latest['EUR_INR'] - previous['EUR_INR']
gbp_change = latest['GBP_INR'] - previous['GBP_INR']
platinum_change = latest['Platinum_1g'] - previous['Platinum_1g']
palladium_change = latest['Palladium_1g'] - previous['Palladium_1g']
copper_change = latest['Copper_1g'] - previous['Copper_1g']
crude_change = latest['Crude_Oil_INR_per_barrel'] - previous['Crude_Oil_INR_per_barrel']
brent_change = latest['Brent_Oil_INR_per_barrel'] - previous['Brent_Oil_INR_per_barrel']
gas_change = latest['Natural_Gas_INR'] - previous['Natural_Gas_INR']


#  USD LIVE 
try:
    usd_live = usd.tail(2).dropna()

    if len(usd_live) < 2:
        raise ValueError("Not enough data")

    usd_price = float(usd_live['Close'].iloc[-1])
    usd_prev = float(usd_live['Close'].iloc[-2])

except:
    usd_price = float(usd['Close'].iloc[-1])
    usd_prev = float(usd['Close'].iloc[-2])


usd_change = usd_price - usd_prev


#  FORMAT 
def format_change(val):
    if val > 0:
        return f"<span style='color:#02ff99; font-weight:bold;'>▲ {abs(val):.2f}</span>"
    elif val < 0:
        return f"<span style='color:#ff4d4d; font-weight:bold;'>▼ {abs(val):.2f}</span>"
    else:
        return "<span style='color:gray;'>0</span>"


#  USD PREDICT 
def predict_usd_next():
    lag1 = usd['Close'].iloc[-1]
    lag2 = usd['Close'].iloc[-2]
    lag3 = usd['Close'].iloc[-3]

    ma3 = usd['Close'].tail(3).mean()
    ma7 = usd['Close'].tail(7).mean()

    X = pd.DataFrame([[lag1, lag2, lag3, ma3, ma7]],
                        columns=['Lag_1','Lag_2','Lag_3','MA_3','MA_7'])

    return usd_model.predict(X)[0]


#  HTML VALUES 
g24_html = format_change(g24_change)
g22_html = format_change(g22_change)
silver_html = format_change(silver_change)
usd_html = format_change(usd_change)
g18_html = format_change(g18_change)
eur_html = format_change(eur_change)
gbp_html = format_change(gbp_change)
platinum_html = format_change(platinum_change)
palladium_html = format_change(palladium_change)
copper_html = format_change(copper_change)
crude_html = format_change(crude_change)
brent_html = format_change(brent_change)
gas_html = format_change(gas_change)

#  SCROLLING TICKER 
ticker_text = f"""
Gold 24K: ₹ {latest['Gold_24K_1g']:.2f} &nbsp;&nbsp;&nbsp;&nbsp;({g24_html})
&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; | &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
Gold 22K: ₹ {latest['Gold_22K_1g']:.2f} &nbsp;&nbsp;&nbsp;&nbsp;({g22_html})
&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; | &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
Gold 18K: ₹ {latest['Gold_18K_1g']:.2f} &nbsp;&nbsp;&nbsp;&nbsp;({g18_html})
&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; | &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
Silver: ₹ {latest['Silver_1g']:.2f} &nbsp;&nbsp;&nbsp;&nbsp;({silver_html})
&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; | &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
Platinum: ₹ {latest['Platinum_1g']:.2f} &nbsp;&nbsp;&nbsp;&nbsp;({platinum_html})
&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; | &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
Palladium: ₹ {latest['Palladium_1g']:.2f} &nbsp;&nbsp;&nbsp;&nbsp;({palladium_html})
&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; | &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
Copper: ₹ {latest['Copper_1g']:.2f} &nbsp;&nbsp;&nbsp;&nbsp;({copper_html})
&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; | &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
USD: ₹ {usd_price:.2f} &nbsp;&nbsp;&nbsp;&nbsp;({usd_html})
&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; | &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
EUR: ₹ {latest['EUR_INR']:.2f} &nbsp;&nbsp;&nbsp;&nbsp;({eur_html})
&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; | &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
GBP: ₹ {latest['GBP_INR']:.2f} &nbsp;&nbsp;&nbsp;&nbsp;({gbp_html})
&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; | &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
Crude Oil: ₹ {latest['Crude_Oil_INR_per_barrel']:.2f} &nbsp;&nbsp;&nbsp;&nbsp;({crude_html})
&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; | &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
Brent Oil: ₹ {latest['Brent_Oil_INR_per_barrel']:.2f} &nbsp;&nbsp;&nbsp;&nbsp;({brent_html})
&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; | &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
Natural Gas: ₹ {latest['Natural_Gas_INR']:.2f} &nbsp;&nbsp;&nbsp;&nbsp;({gas_html})
&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; |
""".replace("\n", " ")

#  RENDER
st.markdown(f"""
<style>
.ticker-container {{
    width: 100%;
    overflow: hidden;
    background: #0e1117;
    padding: 10px 0;
}}

.ticker-track {{
    display: flex;
    width: max-content;
    animation: scroll 25s linear infinite;
}}

.ticker-item {{
    white-space: nowrap;
    color: white;
    font-size: 17px;
    padding-right: 50px;
}}

@keyframes scroll {{
    0% {{ transform: translateX(0); }}
    100% {{ transform: translateX(-50%); }}
}}
</style>

<div class="ticker-container">
    <div class="ticker-track">
        <div class="ticker-item">{ticker_text}</div>
        <div class="ticker-item">{ticker_text}</div>
    </div>
</div>
""", unsafe_allow_html=True)

def create_gold_input(df):

    lag1 = df['Gold_24K_1g'].iloc[-1]
    lag2 = df['Gold_24K_1g'].iloc[-2] if len(df) > 1 else lag1
    lag3 = df['Gold_24K_1g'].iloc[-3] if len(df) > 2 else lag2

    ma7 = df['Gold_24K_1g'].tail(7).mean()
    ma30 = df['Gold_24K_1g'].tail(30).mean()

    usd = df['USD_INR'].iloc[-1]
    usd_change = df['USD_INR'].pct_change().iloc[-1] if len(df) > 1 else 0

    gold22 = df['Gold_22K_1g'].iloc[-1]
    gold18 = df['Gold_18K_1g'].iloc[-1]

    ratio = gold18 / lag1 if lag1 != 0 else 0
    day = df['Date'].iloc[-1].dayofweek

    return pd.DataFrame([[
        lag1, lag2, lag3,
        ma7, ma30,
        usd, usd_change,
        gold22, gold18,
        ratio,
        day
    ]], columns=[
        'Lag_1','Lag_2','Lag_3',
        'MA_7','MA_30',
        'USD_INR','USD_Change',
        'Gold_22K_1g','Gold_18K_1g',
        'Gold_18K_Ratio',
        'DayOfWeek'
    ])
    
def create_silver_input(df):

    lag1 = df['Silver_1g'].iloc[-1]
    lag2 = df['Silver_1g'].iloc[-2] if len(df) > 1 else lag1
    lag3 = df['Silver_1g'].iloc[-3] if len(df) > 2 else lag2

    ma3 = df['Silver_1g'].tail(3).mean()
    ma7 = df['Silver_1g'].tail(7).mean()

    usd = df['USD_INR'].iloc[-1]
    usd_change = df['USD_INR'].pct_change().iloc[-1] if len(df) > 1 else 0

    return pd.DataFrame([[
        lag1, lag2, lag3,
        ma3, ma7,
        usd,
        usd_change
    ]], columns=[
        'Lag_1','Lag_2','Lag_3',
        'MA_3','MA_7',
        'USD_INR',
        'USD_Change'
    ])
    

def create_usd_input(df):

    lag1 = df['USD_INR'].iloc[-1]
    lag2 = df['USD_INR'].iloc[-2] if len(df) > 1 else lag1
    lag3 = df['USD_INR'].iloc[-3] if len(df) > 2 else lag2

    ma3 = df['USD_INR'].tail(3).mean()
    ma7 = df['USD_INR'].tail(7).mean()

    return pd.DataFrame([[lag1, lag2, lag3, ma3, ma7]],
                        columns=['Lag_1','Lag_2','Lag_3','MA_3','MA_7'])
    
def create_eur_input(df):

    lag1 = df['EUR_INR'].iloc[-1]
    lag2 = df['EUR_INR'].iloc[-2] if len(df) > 1 else lag1
    lag3 = df['EUR_INR'].iloc[-3] if len(df) > 2 else lag2

    ma3 = df['EUR_INR'].tail(3).mean()
    ma7 = df['EUR_INR'].tail(7).mean()

    return pd.DataFrame([[lag1, lag2, lag3, ma3, ma7]],
                        columns=['Lag_1','Lag_2','Lag_3','MA_3','MA_7'])
    
def create_gbp_input(df):

    lag1 = df['GBP_INR'].iloc[-1]
    lag2 = df['GBP_INR'].iloc[-2] if len(df) > 1 else lag1
    lag3 = df['GBP_INR'].iloc[-3] if len(df) > 2 else lag2

    ma3 = df['GBP_INR'].tail(3).mean()
    ma7 = df['GBP_INR'].tail(7).mean()

    return pd.DataFrame([[lag1, lag2, lag3, ma3, ma7]],
                        columns=['Lag_1','Lag_2','Lag_3','MA_3','MA_7'])
    
def create_metal_input(df, col):

    lag1 = df[col].iloc[-1]
    lag2 = df[col].iloc[-2] if len(df) > 1 else lag1
    lag3 = df[col].iloc[-3] if len(df) > 2 else lag2

    ma3 = df[col].tail(3).mean()
    ma7 = df[col].tail(7).mean()

    usd = df['USD_INR'].iloc[-1]
    usd_change = df['USD_INR'].pct_change().iloc[-1] if len(df) > 1 else 0

    return pd.DataFrame([[ 
        lag1, lag2, lag3,
        ma3, ma7,
        usd, usd_change
    ]], columns=[
        'Lag_1','Lag_2','Lag_3',
        'MA_3','MA_7',
        'USD_INR','USD_Change'
    ])


def get_prediction(df, metal):

    if metal == "Gold_24K":
        X = create_gold_input(df)
        return gold_model.predict(X)[0]

    elif metal == "Gold_22K":
        X = create_gold_input(df)
        return gold_model.predict(X)[0] * (22/24)

    elif metal == "Gold_18K":
        X = create_gold_input(df)
        return gold_model.predict(X)[0] * (18/24)

    elif metal == "Silver":
        X = create_silver_input(df)
        return silver_model.predict(X)[0]

    elif metal == "USD":
        X = create_usd_input(df)
        return usd_model.predict(X)[0]

    elif metal == "EUR":
        X = create_eur_input(df)
        return eur_model.predict(X)[0]

    elif metal == "GBP":
        X = create_gbp_input(df)
        return gbp_model.predict(X)[0]

    elif metal == "Platinum":
        X = create_metal_input(df, 'Platinum_1g')
        return platinum_model.predict(X)[0]

    elif metal == "Palladium":
        X = create_metal_input(df, 'Palladium_1g')
        return palladium_model.predict(X)[0]

    elif metal == "Copper":
        X = create_metal_input(df, 'Copper_1g')
        return copper_model.predict(X)[0]

    elif metal == "Crude_Oil":
        X = create_metal_input(df, 'Crude_Oil_INR_per_barrel')
        return crude_model.predict(X)[0]

    elif metal == "Brent_Oil":
        X = create_metal_input(df, 'Brent_Oil_INR_per_barrel')
        return brent_model.predict(X)[0]

    elif metal == "Natural_Gas":
        X = create_metal_input(df, 'Natural_Gas_INR')
        return gas_model.predict(X)[0]

    else:
        return None
#  COMMON STYLE 
def render_card(title, value, change=None):
    arrow = ""
    color = "#aaa"

    if change is not None:
        arrow = "▲" if change > 0 else "▼"
        color = "#00ff99" if change > 0 else "#ff4d4d"
        change_html = f"""
        <span style="color:{color}; font-size:16px; margin-left:8px;">
        ({arrow} {abs(change):.2f})
        </span>
        """
    else:
        change_html = ""

    st.markdown(f"""
        <div style="display:flex; flex-direction:column; gap:6px;">
            <span style="color:#aaa; font-size:14px;">{title}</span>
            <span style="font-size:36px; font-weight:400;">₹ {value:.2f} {change_html}</span>
        </div>
        """, unsafe_allow_html=True)
    
#  COMMON UI
def show_section(asset):

    styled_subheader(f"{asset} Overview")
    max_date = df['Date'].max().date()
#  TYPE 
    if asset in ["Gold_24K", "Gold_22K", "Gold_18K", "Silver", "Platinum", "Palladium", "Copper"]:
        asset_type = "metal"
    elif asset in ["USD", "EUR", "GBP"]:
        asset_type = "currency"
    else:
        asset_type = "energy"


#  LAYOUT 
    if asset_type == "currency":
    # FULL WIDTH DATE
        selected_date = st.date_input(
            "Choose Date",
            value=max_date,
            min_value=df['Date'].min().date(),
            max_value=datetime.now().date() + timedelta(days=7),
            key=f"{asset}_date"
        )
    else:
        colA, colB = st.columns(2)

        with colA:
            selected_date = st.date_input(
                "Choose Date",
                value=max_date,
                min_value=df['Date'].min().date(),
                max_value=datetime.now().date() + timedelta(days=7),
                key=f"{asset}_date"
            )


#  WEIGHT / UNIT 
    if asset_type == "metal":
        with colB:
            selected_weight = st.selectbox(
                "Select Weight",
                ["1g", "10g", "100g", "1kg"],
                index=0,
                key=f"{asset}_weight"
            )

        weight_map = {"1g":1, "10g":10, "100g":100, "1kg":1000}
        multiplier = weight_map[selected_weight]
        base_col = f"{asset}_1g"

    elif asset_type == "currency":
        selected_weight = "unit"
        multiplier = 1
        base_col = f"{asset}_INR"

    else:
        with colB:
            selected_weight = st.selectbox(
            "Select Volume",
            ["1 Barrel", "10 Barrel", "100 Barrel", "1000 Barrel"],
            index=0,
            key=f"{asset}_barrel"
            )

        barrel_map = {
            "1 Barrel": 1,
            "10 Barrel": 10,
            "100 Barrel": 100,
            "1000 Barrel": 1000
        }

        multiplier = barrel_map[selected_weight]
        if asset == "Natural_Gas":
            base_col = "Natural_Gas_INR"
        else:
            base_col = f"{asset}_INR_per_barrel"

    #  VALIDATION 
    if base_col not in df.columns:
        st.error(f"{asset} data not available")
        return

    filtered_df = df[df['Date'].dt.date == selected_date]

    #  NORMAL 
    if not filtered_df.empty:
        selected_row = filtered_df.iloc[0]

        base_today = float(selected_row[base_col])
        today_price = base_today * multiplier

        prev_df = df[df['Date'] < pd.to_datetime(selected_date)]
        base_yesterday = float(prev_df.iloc[-1][base_col]) if not prev_df.empty else base_today
        yesterday_price = base_yesterday * multiplier

    #  FUTURE 
    else:
        st.warning("Future date selected — showing prediction")

        temp_df = df.copy()
        days_ahead = (selected_date - max_date).days

        for _ in range(days_ahead):
            next_pred = get_prediction(temp_df, asset)

            new_row = temp_df.iloc[-1].copy()
            new_row['Date'] += timedelta(days=1)
            new_row[base_col] = next_pred  # always 1g

            temp_df = pd.concat([temp_df, pd.DataFrame([new_row])], ignore_index=True)

        selected_row = temp_df.iloc[-1]

        today_price = float(selected_row[base_col]) * multiplier
        yesterday_price = float(temp_df.iloc[-2][base_col]) * multiplier

    change = today_price - yesterday_price

    #  METRICS 
    col1, col2, col3, col4 = st.columns(4)
#  SELECTED 
    with col1:
        render_card("📅 Selected Day", today_price, change)

#  PREVIOUS 
    with col2:
        render_card("🗓️ Previous Day", yesterday_price)

#  HIGHEST 
    with col3:
        render_card("📈 Highest", df[base_col].max() * multiplier)

#  LOWEST 
    with col4:
        render_card("📉 Lowest", df[base_col].min() * multiplier)


#  PREDICTION 
    try:
        prediction_base = get_prediction(df, asset)
        prediction = prediction_base * multiplier

        pred_change = prediction - today_price

        st.markdown("<br>", unsafe_allow_html=True)

        render_card("🎯 Predicted Next Day", prediction, pred_change)

    except Exception as e:
        st.error(f"Prediction error: {e}")
    if asset in ["USD", "EUR", "GBP"]:
        st.caption(f"Selected Date: {selected_date}")
    elif asset_type == "metal":
        st.caption(f"Selected Date: {selected_date} | Weight: {selected_weight}")
    else:
        st.caption(f"Selected Date: {selected_date} | Volume: {selected_weight}")
    #  PRICE TABLE 
    if asset_type in ["metal", "energy"]:
        styled_subheader("Price Table")
        if asset_type == "metal":
            weights = ["1g", "10g", "100g", "1kg"]
            weight_map = {"1g":1, "10g":10, "100g":100, "1kg":1000}
        else:
            weights = ["1 Barrel", "10 Barrel", "100 Barrel", "1000 Barrel"]
            weight_map = {"1 Barrel":1, "10 Barrel":10, "100 Barrel":100, "1000 Barrel":1000}
        rows = []

        for w in weights:
            mul = weight_map[w]

            t = float(selected_row[base_col]) * mul

            prev_df = df[df['Date'] < pd.to_datetime(selected_date)]
            base_prev = float(prev_df.iloc[-1][base_col]) if not prev_df.empty else float(selected_row[base_col])
            y = base_prev * mul

            c = t - y

            change_html = (
                f"<span style='color:#02ff99'>▲ ₹{abs(c):,.2f}</span>"
                if c > 0 else
                f"<span style='color:#ff4d4d'>▼ ₹{abs(c):,.2f}</span>"
            )

            rows.append({
                "Unit": w if asset_type == "metal" else w.replace("g", "Barrel"),
                "Today": f"₹{t:,.2f}",
                "Yesterday": f"₹{y:,.2f}",
                "Change": change_html
            })

        st.markdown(pd.DataFrame(rows).to_html(escape=False, index=False), unsafe_allow_html=True)

    #  FUTURE 7 DAYS 
    styled_subheader("7 Day Prediction")

    future_preds = []
    future_dates = []

    start_date = pd.to_datetime(selected_date)
    temp_df = df.copy()

    if selected_date > max_date:
        temp_df = temp_df.copy()
    for i in range(7):
        next_pred = get_prediction(temp_df, asset)

        future_preds.append(next_pred * multiplier)
        future_dates.append(start_date + timedelta(days=i+1))

        new_row = temp_df.iloc[-1].copy()
        new_row['Date'] += timedelta(days=1)
        new_row[base_col] = next_pred

        temp_df = pd.concat([temp_df, pd.DataFrame([new_row])], ignore_index=True)

    pred_rows = []

    for i in range(7):
        prev_val = today_price if i == 0 else future_preds[i-1]
        curr = future_preds[i]

        diff = curr - prev_val

        change_html = (
            f"<span style='color:#02ff99'>▲ ₹{abs(diff):,.2f}</span>"
            if diff > 0 else
            f"<span style='color:#ff4d4d'>▼ ₹{abs(diff):,.2f}</span>"
        )

        pred_rows.append({
            "Date": future_dates[i].date(),
            "Predicted Price": f"₹{curr:,.2f}",
            "Change": change_html
        })

    st.markdown(pd.DataFrame(pred_rows).to_html(escape=False, index=False), unsafe_allow_html=True)

    #  GRAPH 
    styled_subheader("Price Trend")

    btn_cols = st.columns(6)

    ranges = ["1W", "1M", "3M", "6M", "1Y", "ALL"]

    selected_range = None

    for i, r in enumerate(ranges):
        if btn_cols[i].button(r, key=f"{asset}_btn_{r}"):
            selected_range = r

    if selected_range is None:
        selected_range = "ALL"

    dff = df.copy()

    if selected_range == "1W":
        dff = df.tail(7)
    elif selected_range == "1M":
        dff = df.tail(30)
    elif selected_range == "3M":
        dff = df.tail(90)
    elif selected_range == "6M":
        dff = df.tail(180)
    elif selected_range == "1Y":
        dff = df.tail(365)

    fig = go.Figure()

    fig.add_trace(go.Scatter(
        x=dff['Date'],
        y=dff[base_col] * multiplier,
        mode='lines',
        line=dict(color='#00ff99', width=3),
        name='Actual'
    ))

    fig.add_trace(go.Scatter(
        x=future_dates,
        y=future_preds,
        mode='lines+markers',
        line=dict(color='#ffaa00', width=3, dash='dot'),
        name='Prediction'
    ))

    fig.update_layout(
        template="plotly_dark",
        hovermode="x unified",
        yaxis_title=f"Price ({selected_weight})"
    )

    st.plotly_chart(fig, use_container_width=True)
#  TABS 

tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8, tab9, tab10, tab11, tab12, tab13, tab14 = st.tabs([
    "🪙 Gold 24K",
    "🧈 Gold 22K",
    "🟡 Gold 18K",
    "🔘 Silver",
    "🪨 Platinum",
    "☢️ Palladium",
    "🥮 Copper",
    "💱 USD",
    "💶 EUR",
    "💷 GBP",
    "⛽ Crude Oil",
    "🛢️ Brent Oil",
    "♨️ Natural Gas",
    "🌟 Model Performance"
])


#  TAB 1 
with tab1:
    show_section("Gold_24K")


#  TAB 2 
with tab2:
    show_section("Gold_22K")


#  TAB 3 
with tab3:
    show_section("Gold_18K")


#  TAB 4 
with tab4:
    show_section("Silver")


#  TAB 5 
with tab5:
    show_section("Platinum")


#  TAB 6 
with tab6:
    show_section("Palladium")


#  TAB 7 
with tab7:
    show_section("Copper")


#  TAB 8 
with tab8:
    show_section("USD")


#  TAB 9 
with tab9:
    show_section("EUR")


#  TAB 10 
with tab10:
    show_section("GBP")


#  TAB 11 
with tab11:
    show_section("Crude_Oil")


#  TAB 12 
with tab12:
    show_section("Brent_Oil")


#  TAB 13 
with tab13:
    show_section("Natural_Gas")


#  TAB 14 (FINAL) 
with tab14:

    styled_subheader("🎯 Model Performance (All Assets)")

    rows = []

    def add_row(name, metrics):
        if metrics:
            rows.append({
                "Model": name,
                "MAE": f"{metrics.get('MAE', 0):.2f}",
                "RMSE": f"{metrics.get('RMSE', 0):.2f}",
                "R² Score": f"{metrics.get('R2', 0):.4f}"
            })
    try:
    # REQUIRED
        add_row("Gold", gold_metrics)
        add_row("Silver", silver_metrics)
        add_row("Platinum", platinum_metrics)
        add_row("Palladium", palladium_metrics)
        add_row("Copper", copper_metrics)
        add_row("USD", usd_metrics)
        add_row("EUR", eur_metrics)
        add_row("GBP", gbp_metrics)
        add_row("Crude Oil", crude_metrics)
        add_row("Brent Oil", brent_metrics)
        add_row("Natural Gas", gas_metrics)
    except:
        pass

    table_df = pd.DataFrame(rows)

    #  MEAN R2 
    r2_values = [float(r["R² Score"]) for r in rows]

    if r2_values:
        mean_r2 = sum(r2_values) / len(r2_values)

        table_df.loc[len(table_df)] = {
            "Model": "Model R² Accuracy (Average)",
            "MAE": "-",
            "RMSE": "-",
            "R² Score": f"{mean_r2:.4f}"
        }

    #  TABLE 
    st.markdown(
        table_df.to_html(index=False),
        unsafe_allow_html=True
    )

    st.markdown("---")

    plot_df = table_df[table_df["Model"] != "Model R² Accuracy (Average)"].copy()
    plot_df["R2_float"] = plot_df["R² Score"].astype(float)
    #   METRICS EXPLANATION 
    styled_subheader("📘 Calculation Metrics Explained")

    col1, col2, col3 = st.columns(3)

    with col1:
        st.markdown("""
        ### 📏 MAE (Mean Absolute Error)
        - Average error between actual & predicted
        - Lower = Better

        **Formula:**
        MAE = Σ|Actual - Predicted| / n
        """)

    with col2:
        st.markdown("""
        ### 📐 RMSE (Root Mean Square Error)
        - Penalizes large errors more
        - Sensitive to outliers

        **Formula:**
        RMSE = √(Σ(Actual - Predicted)² / n)
        """)

    with col3:
        st.markdown("""
        ### 🎯 R² Score (Accuracy)
        - Measures model performance
        - 1 = Perfect model
        - 0 = Worst model

        **Formula:**
        R² = 1 - (SS_res / SS_tot)
        """)

    st.markdown("---")

    # BEST MODEL 
    best_model = plot_df.loc[plot_df["R2_float"].idxmax()]
    styled_subheader("🎯 Model Performance (All Assets)")
    #  ANIMATED LINE CHART 
    x_vals = plot_df["Model"].tolist()
    y_vals = plot_df["R2_float"].tolist()

    placeholder = st.empty()
    
    for i in range(1, len(x_vals) + 1):

        fig = go.Figure()

        fig.add_trace(go.Scatter(
            x=x_vals[:i],
            y=y_vals[:i],
            mode='lines+markers',

            marker=dict(
                symbol='circle',
                size=12,
                color='#00a600'
            ),

            line=dict(
                color='#00a600',
                width=3
            ),

            hovertemplate="<b>%{x}</b><br>R²: %{y:.4f}<extra></extra>"
        ))
        
        fig.update_layout(
            xaxis=dict(title="Models"),
            yaxis=dict(title="R² Score"),
            paper_bgcolor="#0E1117",
            plot_bgcolor="#0E1117",
            font=dict(color="white"),
            margin=dict(l=0, r=0, t=40, b=0)
        )

        placeholder.plotly_chart(fig, use_container_width=True)
        time.sleep(0.15)

#  POPUP 
@st.dialog(" ", width="large")
def show_popup(asset, latest):

    icon_map = {
        "Gold": "🧈",
        "Silver": "🔘",
        "Platinum": "🪨",
        "Palladium": "☢️",
        "Copper": "🥮"
    }

    icon = icon_map.get(asset, "⚖️")

    st.markdown(f"## {icon} {asset} Calculator")

    premium_calculator(asset, latest)



def generate_pdf_report(asset, purity, price, qty_grams, base, making_amt, gst_amt, total):

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer)
    styles = getSampleStyleSheet()

    elements = []

    # TITLE
    elements.append(Paragraph(f"<b>{asset} Investment Report</b>", styles['Title']))
    elements.append(Spacer(1, 10))

    elements.append(Paragraph("Financial Market Intelligence", styles['Normal']))
    elements.append(Spacer(1, 20))

    # TABLE
    data = [
        ["Field", "Value"],
        ["Asset", asset],
        ["Purity", purity],
        ["Rate", f"₹{price}"],
        ["Weight", f"{qty_grams} g"],
        ["Total", f"₹{total:,}"]
    ]

    table = Table(data)
    table.setStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.green),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 1, colors.black)
    ])

    elements.append(table)
    elements.append(Spacer(1, 20))

    # DISCLAIMER
    elements.append(Paragraph(
        "<b>⚠ This is an estimated ML-based report. Market risk involved.</b>",
        styles['Normal']
    ))

    doc.build(elements)

    return buffer.getvalue()



def generate_excel_report(asset, purity, price, qty_grams, base, making_amt, gst_amt, total, df):

    #  PREDICTION 
    future_dates, future_prices = [], []
    temp_df = df.copy()
    metal_key = asset if asset != "Gold" else f"Gold_{purity}"

    for i in range(7):
        pred = get_prediction(temp_df, metal_key)
        future_prices.append(int(pred))
        next_date = temp_df['Date'].iloc[-1] + timedelta(days=1)
        future_dates.append(next_date.strftime("%d-%m-%Y"))

        new_row = temp_df.iloc[-1].copy()
        new_row['Date'] = next_date
        if asset == "Gold":
            new_row[f"{metal_key}_1g"] = pred
        else:
            col_map = {"Silver": "Silver_1g", "Platinum": "Platinum_1g",
                       "Palladium": "Palladium_1g", "Copper": "Copper_1g"}
            new_row[col_map[asset]] = pred
        temp_df = pd.concat([temp_df, pd.DataFrame([new_row])], ignore_index=True)

    #  DATA 
    calc_data = [
        ["Asset",           asset],
        ["Purity",          purity if asset == "Gold" else "-"],
        ["Rate (₹/g)",      int(price)],
        ["Weight (g)",      round(qty_grams, 3)],
        ["Base Value",      int(base)],
        ["Making Charges",  int(making_amt)],
        ["GST (+3%)",        int(gst_amt)],
        ["Total Amount",    int(total)],
    ]

    pred_df = pd.DataFrame({
        "Date": future_dates,
        "Predicted Price (₹/g)": future_prices
    })

    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pd.DataFrame(calc_data, columns=["Field", "Value"]).to_excel(writer, index=False, startrow=5, startcol=0)
        pred_df.to_excel(writer, index=False, startrow=5, startcol=3)

        ws = writer.sheets['Sheet1']

        # Helpers─
        def thin_border():
            s = Side(style='thin', color="D1D5DB")
            return Border(left=s, right=s, top=s, bottom=s)

        def set_col_width(col, width):
            ws.column_dimensions[col].width = width

        # Column widths
        set_col_width('A', 22)
        set_col_width('B', 18)
        set_col_width('C', 3)   # spacer
        set_col_width('D', 18)
        set_col_width('E', 22)
        set_col_width('F', 3)   # spacer
        set_col_width('G', 36)  # chart area

        # White background for all used cells 
        white_fill = PatternFill("solid", fgColor="FFFFFF")
        for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=13):
            for cell in row:
                cell.fill = white_fill

        # HEADER BANNER─
        # Deep gold accent bar rows 1-4
        accent_fill = PatternFill("solid", fgColor="1E3A5F")   # deep navy
        gold_fill   = PatternFill("solid", fgColor="C9A84C")   # gold accent

        for row in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=13):
            for cell in row:
                cell.fill = accent_fill

        # Gold accent strip (row 4)
        for col in range(1, 14):
            ws.cell(row=4, column=col).fill = gold_fill

        ws.row_dimensions[1].height = 14
        ws.row_dimensions[2].height = 28
        ws.row_dimensions[3].height = 22
        ws.row_dimensions[4].height = 5

        # Title
        ws.merge_cells("A2:M2")
        ws["A2"] = f"{asset} Investment Report"
        ws["A2"].font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

        # Subtitle
        ws.merge_cells("A3:M3")
        ws["A3"] = "Financial Market Intelligence  •  Predictive Analytics"
        ws["A3"].font = Font(name="Calibri", size=10, italic=True, color="B0C4DE")
        ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

        # SECTION LABELS
        ws.row_dimensions[5].height = 20

        def section_label(cell_ref, text):
            c = ws[cell_ref]
            c.value = text
            c.font = Font(name="Calibri", size=9, bold=True, color="6B7280")
            c.alignment = Alignment(horizontal="left", vertical="center")

        section_label("A5", "INVESTMENT DETAILS")
        section_label("D5", "7-DAY PRICE FORECAST")

        # TABLE HEADERS
        hdr_fill = PatternFill("solid", fgColor="1E3A5F")
        hdr_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        hdr_align = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[6].height = 22

        for col_letter, label in [("A", "Field"), ("B", "Value")]:
            c = ws[f"{col_letter}6"]
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = hdr_align
            c.border = thin_border()

        for col_letter, label in [("D", "Date"), ("E", "Predicted Price (₹/g)")]:
            c = ws[f"{col_letter}6"]
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = hdr_align
            c.border = thin_border()

        # TABLE ROWS
        alt_fill  = PatternFill("solid", fgColor="F0F4F8")
        norm_fill = PatternFill("solid", fgColor="FFFFFF")
        row_font  = Font(name="Calibri", size=10, color="1F2937")
        val_font  = Font(name="Calibri", size=10, bold=True, color="1E3A5F")

        highlight_rows = {13}   # Total Amount row (row index 7 in calc_data = excel row 13)

        for i in range(8):
            excel_row = 7 + i
            ws.row_dimensions[excel_row].height = 20
            fill = alt_fill if i % 2 == 0 else norm_fill

            # Highlight total row
            if excel_row == 13:
                fill = PatternFill("solid", fgColor="EFF6FF")

            for col_letter in ["A", "B"]:
                c = ws[f"{col_letter}{excel_row}"]
                c.fill = fill
                c.border = thin_border()
                if col_letter == "A":
                    c.font = row_font
                    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                else:
                    c.font = val_font if excel_row == 13 else row_font
                    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        # Bold the total row label too
        ws["A13"].font = Font(name="Calibri", size=10, bold=True, color="1E3A5F")

        # Prediction table rows
        for i in range(7):
            excel_row = 7 + i
            ws.row_dimensions[excel_row].height = 20
            fill = alt_fill if i % 2 == 0 else norm_fill
            for col_letter in ["D", "E"]:
                c = ws[f"{col_letter}{excel_row}"]
                c.fill = fill
                c.border = thin_border()
                c.font = row_font
                c.alignment = Alignment(
                    horizontal="center" if col_letter == "D" else "right",
                    vertical="center",
                    indent=1
                )

        # CHART (3D Bar)─
        chart = BarChart3D()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.title = "7-Day Price Forecast (₹/g)"
        chart.y_axis.title = "Price (₹/g)"
        chart.x_axis.title = "Date"
        chart.style = 26
        chart.width  = 16
        chart.height = 10

        # Data from prediction table (col E = 5, rows 6-13 with header)
        data = Reference(ws, min_col=5, min_row=6, max_row=13)
        cats = Reference(ws, min_col=4, min_row=7, max_row=13)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Place chart to the right of both tables
        ws.add_chart(chart, "G5")

        # DISCLAIMER─
        ws.row_dimensions[22].height = 22
        ws.merge_cells("A22:M22")
        ws["A22"] = "⚠  DISCLAIMER: Market predictions are indicative only. Past performance does not guarantee future results. Please consult a financial advisor before investing."
        ws["A22"].font = Font(name="Calibri", size=9, italic=True, color="9CA3AF")
        ws["A22"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws["A22"].fill = PatternFill("solid", fgColor="F9FAFB")
        disc_border = Border(
            top=Side(style='thin', color="E5E7EB"),
            bottom=Side(style='thin', color="E5E7EB")
        )
        ws["A22"].border = disc_border

        # FOOTER
        footer_fill = PatternFill("solid", fgColor="1E3A5F")
        gold_fill2  = PatternFill("solid", fgColor="C9A84C")

        ws.row_dimensions[23].height = 4
        for col in range(1, 14):
            ws.cell(row=23, column=col).fill = gold_fill2

        ws.row_dimensions[24].height = 24
        for col in range(1, 14):
            ws.cell(row=24, column=col).fill = footer_fill

        ws.merge_cells("A24:M24")
        ws["A24"] = "© 2026  •  Developed by Daksh Vasani  |  Advanced Analytics  •  Machine Learning  •  Financial Insights"
        ws["A24"].font = Font(name="Calibri", size=10, italic=True, color="B0C4DE")
        ws["A24"].alignment = Alignment(horizontal="center", vertical="center")

        # LOGO (optional)
        try:
            logo = Image("logo.png")
            logo.width  = 90
            logo.height = 40
            ws.add_image(logo, "L2")
        except Exception:
            pass

        # Freeze panes
        ws.freeze_panes = "A7"

        # Print settings
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToPage   = True
        ws.page_setup.fitToWidth  = 1
        ws.sheet_view.showGridLines = False

    return output.getvalue()
#  MAIN 

def premium_calculator(asset, latest):

    price_map = {
        "Gold": latest["Gold_24K_1g"],
        "Silver": latest["Silver_1g"],
        "Platinum": latest["Platinum_1g"],
        "Palladium": latest["Palladium_1g"],
        "Copper": latest["Copper_1g"]
    }

    prefix = f"popup_{asset}"

    #  STATE 
    st.session_state.setdefault(f"{prefix}_mode", "weight")
    st.session_state.setdefault(f"{prefix}_purity", "24K")

    st.session_state.setdefault(f"{prefix}_val", 10)
    st.session_state.setdefault(f"{prefix}_val_amt", 10000)
    st.session_state.setdefault(f"{prefix}_unit", "Gram")

    left, right = st.columns([1, 1.2])

    #  LEFT 
    with left:

        c1, c2 = st.columns(2)

        if c1.button("Calculate By Weight", key=f"{prefix}_w"):
            st.session_state[f"{prefix}_mode"] = "weight"

        if c2.button("Calculate By Amount", key=f"{prefix}_a"):
            st.session_state[f"{prefix}_mode"] = "amount"

        mode = st.session_state[f"{prefix}_mode"]
        purity = st.session_state[f"{prefix}_purity"]

        #  GOLD 
        if asset == "Gold":
            st.markdown("### Gold Purity")
            p1, p2, p3 = st.columns(3)

            if p1.button("24K"): st.session_state[f"{prefix}_purity"] = "24K"
            if p2.button("22K"): st.session_state[f"{prefix}_purity"] = "22K"
            if p3.button("18K"): st.session_state[f"{prefix}_purity"] = "18K"

        purity = st.session_state[f"{prefix}_purity"]

        #  PRICE 
        base_price = price_map[asset]

        if asset == "Gold":
            purity_map = {"24K":1, "22K":22/24, "18K":18/24}
            price = base_price * purity_map.get(purity, 1)
        else:
            price = base_price

        #  UNIT 
        unit_map = {
            "Gram": 1,
            "Sovereign / Pavan": 8,
            "Tola": 11.664,
            "KG": 1000
        }

        #  MODE: WEIGHT 
        if mode == "weight":

            q1, q2 = st.columns([3, 1])

            with q2:
                unit = st.selectbox(
                    "Unit",
                    ["Gram", "Sovereign / Pavan", "Tola", "KG"],
                    key=f"{prefix}_unit"
                )

            multiplier = unit_map[unit]
            max_val = 1000 if unit == "Gram" else 100

            if f"{prefix}_input" not in st.session_state:
                st.session_state[f"{prefix}_input"] = st.session_state[f"{prefix}_val"]

            if f"{prefix}_slider" not in st.session_state:
                st.session_state[f"{prefix}_slider"] = st.session_state[f"{prefix}_val"]

            st.session_state[f"{prefix}_val"] = min(st.session_state[f"{prefix}_val"], max_val)
            st.session_state[f"{prefix}_input"] = min(st.session_state[f"{prefix}_input"], max_val)
            st.session_state[f"{prefix}_slider"] = min(st.session_state[f"{prefix}_slider"], max_val)

            if st.session_state[f"{prefix}_input"] != st.session_state[f"{prefix}_val"]:
                st.session_state[f"{prefix}_val"] = st.session_state[f"{prefix}_input"]
            elif st.session_state[f"{prefix}_slider"] != st.session_state[f"{prefix}_val"]:
                st.session_state[f"{prefix}_val"] = st.session_state[f"{prefix}_slider"]

            st.session_state[f"{prefix}_input"] = st.session_state[f"{prefix}_val"]
            st.session_state[f"{prefix}_slider"] = st.session_state[f"{prefix}_val"]

            with q1:
                st.number_input("Quantity", 1, max_val, key=f"{prefix}_input")

            st.slider(" ", 1, max_val, key=f"{prefix}_slider")

            qty = float(st.session_state[f"{prefix}_val"])

            #LOCK VALUE (UI = exact input)
            final_weight = qty

            qty_grams = qty * multiplier

            making = st.slider("Making Charge (%)", 0, 50, 10)
            gst = st.checkbox("Include GST (3%)", True)

            base = qty_grams * price

        #  MODE: AMOUNT 
        else:

            if f"{prefix}_amt_input" not in st.session_state:
                st.session_state[f"{prefix}_amt_input"] = st.session_state[f"{prefix}_val_amt"]
            if f"{prefix}_amt_slider" not in st.session_state:
                st.session_state[f"{prefix}_amt_slider"] = st.session_state[f"{prefix}_val_amt"]

            if st.session_state[f"{prefix}_amt_input"] != st.session_state[f"{prefix}_val_amt"]:
                st.session_state[f"{prefix}_val_amt"] = st.session_state[f"{prefix}_amt_input"]
            elif st.session_state[f"{prefix}_amt_slider"] != st.session_state[f"{prefix}_val_amt"]:
                st.session_state[f"{prefix}_val_amt"] = st.session_state[f"{prefix}_amt_slider"]

            st.session_state[f"{prefix}_amt_input"] = st.session_state[f"{prefix}_val_amt"]
            st.session_state[f"{prefix}_amt_slider"] = st.session_state[f"{prefix}_val_amt"]

            st.number_input("Amount (₹)", 0, 1000000, key=f"{prefix}_amt_input")
            st.slider("Amount", 0, 1000000, key=f"{prefix}_amt_slider")

            amount = float(st.session_state[f"{prefix}_val_amt"])
            original_amount = amount

            making = st.slider("Making Charge (%)", 0, 50, 10)
            gst = st.checkbox("Include GST (3%)", True)

            base = amount

            qty_grams = base / price if price > 0 else 0

            # STABLE VALUE (no drift)
            final_weight = float(f"{qty_grams:.3f}")

        #  FINAL CALC 
        making_amt = base * (making/100)
        subtotal = base + making_amt
        gst_amt = subtotal * 0.03 if gst else 0
        total = subtotal + gst_amt

        #  DOWNLOAD 
        st.markdown("<br><br>", unsafe_allow_html=True)

        excel_data = generate_excel_report(
            asset, purity, price, qty_grams,
            base, making_amt, gst_amt, total, df
        )

        st.download_button(
            label="📥 Download Full Report",
            data=excel_data,
            file_name=f"{asset}_full_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    #  RIGHT 
    with right:

        placeholder = st.empty()

        for i in range(26):

            if mode == "amount":
                val = final_weight * (i/25)
                text = f"{val:.3f} g"
                title = f"{asset} Weight"
            else:
                val = int(total * (i/25))
                text = f"₹ {val:,}"
                title = "Total Amount"

            placeholder.markdown(f"""
                <div style="background:#16a34a;padding:18px;border-radius:14px;color:white">
                    <div style="font-size:25px;">{title}</div>
                    <div style="font-size:45px;font-weight:800;">{text}</div>
                </div>""", unsafe_allow_html=True)

            time.sleep(0.008)
        #  CURRENT RATE CARD (GOODRETURNS STYLE) 
        rate_label = f"{purity} Current Rate" if asset == "Gold" else f"{asset} Current Rate"

        st.markdown(f"""
        <div style="background:rgba(255,255,255,0.12);padding:14px 16px;border-radius:12px;color:white;margin-top:12px;">
            <div style="font-size:14px;opacity:0.85;margin:0;line-height:1.2;">{rate_label}</div>
            <div style="font-size:18px;font-weight:700;margin-top:4px;line-height:1.2;">₹{int(price):,}/gram</div>
        </div>
        """, unsafe_allow_html=True)

        #  BREAKDOWN 
        st.markdown("### Calculation Breakdown")

        def row(label, val, money=True):
            txt = f"₹ {int(val):,}" if money else f"{val:.3f} g"
            st.markdown(f"""
            <div style="display:flex;justify-content:space-between;padding:6px 0;">
                <span>{label}</span>
                <b>{txt}</b>
            </div>
            """, unsafe_allow_html=True)

        if mode == "amount":
            row("Amount Entered", original_amount)
        else:
            row("Weight", final_weight, False)

        row("Base Value", base)
        row("Making Charges", making_amt)
        st.markdown("<hr style='border-color:#444'>", unsafe_allow_html=True)
        row("Subtotal", subtotal)
        row("GST (3%)", gst_amt)

        st.markdown(f"""
        <div style="margin-top:15px;background:#15803d;padding:12px;border-radius:12px;color:white;font-weight:bold;display:flex;justify-content:space-between;">
            <span>{"Final Weight" if mode=="amount" else "Total Amount"}</span>
            <span>{"{:.3f} g".format(final_weight) if mode=="amount" else "₹ {:,}".format(int(total))}</span>
        </div>
        """, unsafe_allow_html=True)
#  SIDEBAR 
with st.sidebar:
    st.markdown("""
    <style>
    [data-testid="stSidebar"] [data-testid="stElementContainer"] {
        width: 100% !important;
    }

    [data-testid="stSidebar"] button {
        width: 100% !important;
        height: 55px !important;
        border-radius: 12px !important;
        font-size: 16px !important;
        font-weight: 600 !important;
        text-align: left !important;
        padding-left: 15px !important;
        display: flex !important;
        align-items: center !important;
        justify-content: flex-start !important;
    }
    
    [data-testid="stSidebar"] button div p {
        width: 100% !important;
        text-align: left !important;
    }
    </style>
    """, unsafe_allow_html=True)


    # Professional Header Structure
    st.markdown("""
        <h2 style='
            color:white;
            border-left:6px solid #00a600;
            padding-left:12px;
            font-weight:bold;'header-text">⚖️ Metal Calculator</h2>
        """, unsafe_allow_html=True)
    st.markdown("<div style='height:20px;'></div>", unsafe_allow_html=True)

    if st.button("Gold"):
        st.session_state.popup = "Gold"

    if st.button("Silver"):
        st.session_state.popup = "Silver"

    if st.button("Platinum"):
        st.session_state.popup = "Platinum"

    if st.button("Palladium"):
        st.session_state.popup = "Palladium"

    if st.button("Copper"):
        st.session_state.popup = "Copper"


#  TRIGGER 
if "popup" not in st.session_state:
    st.session_state.popup = None

if st.session_state.popup:
    asset_to_show = st.session_state.popup
    st.session_state.popup = None
    show_popup(asset_to_show, latest)

#  FOOTER
st.markdown("---")
st.caption("© 2026 • Developed by Daksh Vasani | Advanced Analytics • Machine Learning • Financial Insights")
